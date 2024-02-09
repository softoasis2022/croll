import time
import requests

from bs4 import BeautifulSoup
import bs4
from bs4 import BeautifulSoup as bs

from selenium import webdriver
from selenium import common
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

from selenium import webdriver 
import chromedriver_autoinstaller
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import warnings

from openpyxl import Workbook
from openpyxl import load_workbook

from datetime import datetime
import os



def write():
    # 엑셀파일 쓰기
    write_wb = Workbook()

    # 이름이 있는 시트를 생성
    write_ws = write_wb.create_sheet('생성시트')

    # Sheet1에다 입력
    write_ws = write_wb.active
    #write_ws['A1'] = '숫자'

    #행 단위로 추가
    write_ws.append([1,2,3])

    #셀 단위로 추가
    #write_ws.cell(5, 5, '5행5열')
    write_wb.save("D://시간.xlsx")

def read():
    # data_only=True로 해줘야 수식이 아닌 값으로 받아온다. 
        load_wb = load_workbook("D://숫자.xlsx", data_only=True)
        # 시트 이름으로 불러오기 
        load_ws = load_wb['Sheet1']

        # 셀 주소로 값 출력
        print(load_ws['B2'].value)

        # 셀 좌표로 값 출력
        print(load_ws.cell(3, 2).value)


        # 지정한 셀의 값 출력

        get_cells = load_ws['B3' : 'B6']
        for row in get_cells:
            for cell in row:
                print(cell.value)

        # 모든 행 단위로 출력

        for row in load_ws.rows:
            print(row)

        # 모든 열 단위로 출력

        for column in load_ws.columns:
            print(column)

        # 모든 행과 열 출력

        all_values = []
        for row in load_ws.rows:
            row_value = []
            for cell in row:
                row_value.append(cell.value)
            all_values.append(row_value)
        print(all_values)

        load_ws.cell(3, 3, 51470)
        load_ws.cell(4, 3, 21470)
        load_ws.cell(5, 3, 1470)
        load_ws.cell(6, 3, 6470)
        load_wb.save("C:/Users/Administrator/Desktop/기준/프로그래밍/과제대행/주식데이터크롤링/output.xlsx")
now = datetime.now()
def tm():
    
    print("현재 : ", now)
    print("현재 날짜 : ", now.date())
    print("현재 시간 : ", now.time())
    print("timestamp : ", now.timestamp())
    print("년 : ", now.year)
    print("월 : ", now.month)
    print("일 : ", now.day)
    print("시 : ", now.hour)
    print("분 : ", now.minute)
    print("초 : ", now.second)
    print("마이크로초 : ", now.microsecond)
    print("요일 : ", now.weekday())
    print("문자열 변환 : ", now.strftime('%Y-%m-%d %H:%M:%S'))
    return now.strftime('%Y-%m-%d %H:%M:%S')


def makedirs(path):
    if not os.path.exists(path):
        os.makedirs(path)

makedirs("D://"+"사진"+"/인디")
write()